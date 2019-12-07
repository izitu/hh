# Отдельно взятый скрипт - выкачиваем вакансии отдельновзятой фирмы (в данном случае - Газпром)

import openpyxl, json

# проверяем словарь - если есть в словаре стоп слово - вакансию не печатаем
def dict(name):
	#stopw = []
	stopw = ['АЗС','охр','Переводчик','Лаборант','питания', 'Химки','Электромонтер','1С','1C', 'Битрикс','Механик','Электромонтажник','Водитель','ТРИЗ',
			 'логистике','технолог','продаж','Токарь','юрист','Клиент-менеджер', 'Эксплуатации','Оператор','HRBP','документообороту','простоя',
			 'Консультант', 'планово-экономического','сметной','Слесарь','делопроизводству', 'РЗА','задолженностью','контроллинга','Стажер',
            'Бетонщик','Кровельщик','аудитор','Помощник','Маляр','сметчик','Мерчендайзер','Владимир','Финансовый','логист','Логист','Бухгалтер',
			 'геолог','Продавец','Ассистент','ассистент','шеф','тренер','Повар','юрисконсульт','налогового','языку','водитель','Супервайзер','персоналу','Тамбов','Бармен',
			 'сестра','Врач','котельной','юрист','Юрист','льготам','персонала','животноводству','Агроном','сельхозтехнике','агроном','акушер']
	for w in stopw:
		if name.count(w)!=0:
			return False
	return True

# одна фирма например Газпром id 39305
onefirm = '39305'
firm_name = 'Газпром'
# создаем новый excel-файл
wb2 = openpyxl.Workbook()

# добавляем новый лист
wb2.create_sheet(title='Первый лист', index=0)
# получаем лист, с которым будем работать
sheet2 = wb2['Первый лист']
# заголовок
cell2 = sheet2.cell(row=1, column=1)
cell2.value = 'Фирма'
cell2 = sheet2.cell(row=1, column=2)
cell2.value = 'Вакансия'
cell2 = sheet2.cell(row=1, column=3)
cell2.value = 'Описание'
cell2 = sheet2.cell(row=1, column=4)
cell2.value = 'Требование'
cell2 = sheet2.cell(row=1, column=5)
cell2.value = 'URL'
cell2 = sheet2.cell(row=1, column=6)
cell2.value = 'от'
cell2 = sheet2.cell(row=1, column=7)
cell2.value = 'до'
cell2 = sheet2.cell(row=1, column=8)
cell2.value = 'дата'

rowi = 2
try:
	with open(onefirm + '.api', 'r', encoding='utf-8') as fh:
		data = json.load(fh)
		#print(data)
except:
	print('нет вакансий')
	pass

for vac in range(0,len(data['items'])):
	if dict(data['items'][vac]['name']):
		if (data['items'][vac]['salary']) == None:
			fr = ''
			to = ''
		else:
			fr = data['items'][vac]['salary']['from']
			to = data['items'][vac]['salary']['to']
				#print(firm_name, data['items'][vac])#['name']['alternate_url']
		tt = data['items'][vac]['published_at'][:10].split('-')
		dt = tt[2]+'.'+tt[1]+'.'+tt[0]
		print(rowi, firm_name, data['items'][vac]['name'], "|",
					data['items'][vac]['snippet']['responsibility'], "|",data['items'][vac]['snippet']['requirement'], "|",
					data['items'][vac]['alternate_url'], "|", fr, to, "|", dt)
		cell2 = sheet2.cell(row=rowi, column=1)
		cell2.value = firm_name
		cell2 = sheet2.cell(row=rowi, column=2)
		cell2.value = data['items'][vac]['name']
		cell2 = sheet2.cell(row=rowi, column=3)
		cell2.value = data['items'][vac]['snippet']['responsibility']
		cell2 = sheet2.cell(row=rowi, column=4)
		cell2.value = data['items'][vac]['snippet']['requirement']
		cell2 = sheet2.cell(row=rowi, column=5)
		cell2.value = data['items'][vac]['alternate_url']
		cell2 = sheet2.cell(row=rowi, column=6)
		cell2.value = fr
		cell2 = sheet2.cell(row=rowi, column=7)
		cell2.value = to
		cell2 = sheet2.cell(row=rowi, column=8)
		cell2.value = dt
		rowi = rowi + 1
wb2.save(onefirm +'.xlsx')

exit(0)
#загрузить из json
#with open('4023.api', 'r', encoding='utf-8') as fh:
#	data = json.load(fh)
#try:
#	print(data['items'][0])
#except:
#	pass
#смотрим все апи файлы
files = os.listdir('.')
apis = filter(lambda x: x.endswith('.api'), files)
for api in apis:
	print(api)
exit(0)