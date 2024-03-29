# запускается ТРЕТЬИМ - разбираем файлы - заливаем в эксель
# БАГ!!! - ИЗБАВИТЬСЯ ОТ ДУБЛИРОВАНИЯ

import json
import openpyxl


# проверяем словарь - если есть в словаре стоп слово - вакансию не печатаем dict
def dict_iskl(name):
    with open('stop-dict.json', encoding='utf-8') as data_file:
        stopw = json.load(data_file, strict=False)
    for w in stopw:
        if name.count(w) != 0:
            return False
    return True


# читаем excel-файл
wb = openpyxl.load_workbook('100bighh/firms.xlsx')

# печатаем список листов
sheets = wb.sheetnames
for sheet in sheets:
    print(sheet)

# получаем активный лист
sheet = wb.active
rows = sheet.max_row
print(rows)

# создаем новый excel-файл
wb2 = openpyxl.Workbook()

# добавляем новый лист
wb2.create_sheet(title='Первый лист', index=0)
# получаем лист, с которым будем работать
sheet2 = wb2['Первый лист']
# заголовок
cell2 = sheet2.cell(row=1, column=1)
cell2.value = 'Фирма'
cell2 = sheet2.cell(row=1, column=2)
cell2.value = 'Вакансия'
cell2 = sheet2.cell(row=1, column=3)
cell2.value = 'Описание'
cell2 = sheet2.cell(row=1, column=4)
cell2.value = 'Требование'
cell2 = sheet2.cell(row=1, column=5)
cell2.value = 'URL'
cell2 = sheet2.cell(row=1, column=6)
cell2.value = 'от'
cell2 = sheet2.cell(row=1, column=7)
cell2.value = 'до'
cell2 = sheet2.cell(row=1, column=8)
cell2.value = 'дата'

rowi = 2
for i in range(2, rows + 1):
    cell = sheet.cell(row=i, column=1)
    firm_no = str(cell.value)
    cell = sheet.cell(row=i, column=2)
    firm_name = str(cell.value)
    cell = sheet.cell(row=i, column=3)
    firm_id = str(cell.value)
    print(firm_no, firm_name, firm_id, firm_id + '.api')
    try:
        with open("100bighh/" + firm_id + ".api", "r", encoding="utf-8") as fh:
            data = json.load(fh)
        # print(data)
    except:
        print('нет вакансий')
        pass

    for vac in range(0, len(data['items'])):
        if dict_iskl(data['items'][vac]['name']):
            if (data['items'][vac]['salary']) is None:
                fr = ''
                to = ''
            else:
                fr = data['items'][vac]['salary']['from']
                to = data['items'][vac]['salary']['to']
            # print(firm_name, data['items'][vac])#['name']['alternate_url']
            tt = data['items'][vac]['published_at'][:10].split('-')
            dt = tt[2] + '.' + tt[1] + '.' + tt[0]
            print(rowi, firm_name, data['items'][vac]['name'], "|",
                  data['items'][vac]['snippet']['responsibility'], "|", data['items'][vac]['snippet']['requirement'],
                  "|",
                  data['items'][vac]['alternate_url'], "|", fr, to, "|", dt)
            cell2 = sheet2.cell(row=rowi, column=1)
            cell2.value = firm_name
            cell2 = sheet2.cell(row=rowi, column=2)
            cell2.value = data['items'][vac]['name']
            cell2 = sheet2.cell(row=rowi, column=3)
            cell2.value = data['items'][vac]['snippet']['responsibility']
            cell2 = sheet2.cell(row=rowi, column=4)
            cell2.value = data['items'][vac]['snippet']['requirement']
            cell2 = sheet2.cell(row=rowi, column=5)
            cell2.value = data['items'][vac]['alternate_url']
            cell2 = sheet2.cell(row=rowi, column=6)
            cell2.value = fr
            cell2 = sheet2.cell(row=rowi, column=7)
            cell2.value = to
            cell2 = sheet2.cell(row=rowi, column=8)
            cell2.value = dt
            rowi = rowi + 1
wb2.save("100bighh/all-vac.xlsx")
