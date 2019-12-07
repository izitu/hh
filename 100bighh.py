# запускается ПЕРВЫМ
# выкачиваем информацию о лучших фирмах из статьи
# https://hh.ru/article/303400, собираем номер - название и id

import requests, bs4
import openpyxl

def webtofile_100firm():
	# сохраняем в файл нашу страницу

	base_url = 'https://hh.ru/article/303400'
	headers = {'accept':'*/*',
			'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:70.0) Gecko/20100101 Firefox/70.0'}
	session = requests.Session()
	s0 = session.get(base_url,headers=headers)
	#cs0=requests.get('https://hh.ru/article/303400')
	b=bs4.BeautifulSoup(s0.text, "html.parser")

	#f_all_row = b.findall(tr)
	# режим экономии
	s_to_file = open("100firm_raw.html", "w", encoding="utf-8")
	s_to_file.write(str(b))
	s_to_file.close()
	print(b)



# вызываем функцию скачивания - в идеале она должна всега скачиваться
# webtofile_100firm()
# читаем из файла
open_raw = open("100firm_raw.html", "r", encoding="utf-8")
s1 = open_raw.read()
b=bs4.BeautifulSoup(s1, "html.parser")
open_raw.close()
# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title = 'Первый лист', index = 0)
# получаем лист, с которым будем работать
sheet = wb['Первый лист']
# заголовок
cell = sheet.cell(row=1, column=1)
cell.value = 'No'
cell = sheet.cell(row=1, column=2)
cell.value = 'Название'
cell = sheet.cell(row=1, column=3)
cell.value = 'ID'
cell = sheet.cell(row=1, column=4)
cell.value = 'Описание'
cell = sheet.cell(row=1, column=5)
cell.value = 'URL'

rows = b.select('.cms-table__body-row')
r = 2
for row in rows:
	try:
		firm_no = row.select('.cms-table__body-cell')[0].text.strip()
		firm_name = row.select('.cms-table__body-cell')[1].text.strip()
		firm_spec = row.select('.cms-table__body-cell')[2].text.strip()
		firm_id = row.select('.cms-table__body-cell')[1].select('a')[0].get('href').split('/')[-1].split('?')[0]
		firm_url = row.select('.cms-table__body-cell')[1].select('a')[0].get('href')
		#print(firm_no, firm_name, firm_spec, firm_url, firm_id)
		cell = sheet.cell(row=r, column=1)
		cell.value = firm_no
		cell = sheet.cell(row=r, column=2)
		cell.value = firm_name
		cell = sheet.cell(row=r, column=3)
		cell.value = firm_id
		cell = sheet.cell(row=r, column=4)
		cell.value = firm_spec
		cell = sheet.cell(row=r, column=5)
		cell.value = firm_url
		r = r + 1
	except:
		pass
# Добавим ОЭК и МОЭСК
cell = sheet.cell(row=r, column=1)
cell.value = str(int(firm_no)+1)
cell = sheet.cell(row=r, column=2)
cell.value = 'ПАО ОЭК'
cell = sheet.cell(row=r, column=3)
cell.value = '102401'
cell = sheet.cell(row=r+1, column=1)
cell.value = str(int(firm_no)+2)
cell = sheet.cell(row=r+1, column=2)
cell.value = 'ПАО «МОЭК»'
cell = sheet.cell(row=r+1, column=3)
cell.value = '58138'

wb.save('firms.xlsx')

# читаем excel-файл
wb = openpyxl.load_workbook('firms.xlsx')

# печатаем список листов
sheets = wb.sheetnames
for sheet in sheets:
	print(sheet)

# получаем активный лист
sheet = wb.active
rows = sheet.max_row
print(rows)

for i in range(2, rows + 1):
	cell = sheet.cell(row=i, column=1)
	firm_no = str(cell.value)
	cell = sheet.cell(row=i, column=2)
	firm_name = str(cell.value)
	cell = sheet.cell(row=i, column=3)
	firm_id = str(cell.value)
	print(firm_no, firm_name, firm_id)
exit(0)
# берем все ссылки с фирмами
f_all = b.find_all('a', attrs='cms-link')
# удаляем те первые, не нужные
del f_all[0:2]
my_file = open("firm.txt", "w")


for rec in f_all:
	p_name=rec.text
	p_url=rec.get('href').split('/')[-1].split('?')
	print(p_name, p_url[0])
	my_file.write(p_name+' '+p_url[0]+'\n')

my_file.write('ПАО ОЭК'+' '+'102401'+'\n')
my_file.write('ПАО «МОЭК»'+' '+'58138'+'\n')
my_file.close()